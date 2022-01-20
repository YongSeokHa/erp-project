import pymssql
from openpyxl import load_workbook
from collections import deque
import heapq
import datetime
from address import *


class Target_num_code:
    conn = pymssql.connect(host=hq_host, user=id, password=pw, database='HQERP',
                           charset='utf8')
    cursor = conn.cursor()

    def __init__(self, num):
        self.num = num

    def find_target_number_codes(self, position):
        self.position = position
        if self.position == "대":
            self.cursor.execute("SELECT ITEM_CD, ITEM_NM, SPEC, ITEM_ACCT "
                                "FROM B_ITEM "
                                f"WHERE ITEM_CD LIKE '{str(self.num)}%'")

            row = self.cursor.fetchall()
            # 코드, 품목명, 규격, acct
            return row

class Code(Target_num_code):
    conn = pymssql.connect(host=hq_host, user=id, password=pw, database='HQERP',
                           charset='utf8')
    cursor = conn.cursor()

    client_list = {}


    def __init__(self, code):
        self.code = code
        # self.num = int(str(self.code)[0:2])

    # 원자재 모델 찾기 return [[prnt_item_cd,item_nm,valid_from_dt]]
    def find_models(self):
        self.cursor.execute("SELECT D.PRNT_ITEM_CD, D.CHILD_ITEM_CD, I.ITEM_NM, I.SPEC, I.VALID_FROM_DT, I.ITEM_ACCT "
                       "FROM P_BOM_DETAIL D "
                       "JOIN B_ITEM I ON D.CHILD_ITEM_CD = I.ITEM_CD "
                       f"WHERE D.CHILD_ITEM_CD = '{str(self.code)}'AND D.PRNT_PLANT_CD = 'P1'")

        row = self.cursor.fetchall()
        finding_models = deque(row)
        models = []

        # finding_models = [prnt_item_cd, child_item_cd, item_nm, spec, acct, valid_from_dt]
        while finding_models:
            model = finding_models.popleft()
            prnt_item_cd = model[0]
            child_item_cd = model[1]
            item_nm = model[2]
            spec = model[3]
            valid_from_dt = model[4]
            acct = model[5]

            if acct == '10':
                models.append([prnt_item_cd,item_nm,valid_from_dt])
                continue

            self.cursor.execute("SELECT D.PRNT_ITEM_CD, D.CHILD_ITEM_CD, I.ITEM_NM, I.SPEC, I.VALID_FROM_DT, I.ITEM_ACCT "
                           "FROM P_BOM_DETAIL D "
                           "JOIN B_ITEM I ON D.PRNT_ITEM_CD = I.ITEM_CD "
                           f"WHERE '{prnt_item_cd}' = D.CHILD_ITEM_CD")

            row = self.cursor.fetchone()
            while row:
                finding_models.append(row)
                row = self.cursor.fetchone()

        return models # [prnt_item_cd,item_nm,valid_from_dt]

    # 원자재 revison 최신 모델 찾기 return [[item_nm]]
    def find_models_recently_revison(self):
        self.cursor.execute("SELECT D.PRNT_ITEM_CD, D.CHILD_ITEM_CD, I.ITEM_NM, I.SPEC, I.VALID_FROM_DT, I.ITEM_ACCT "
                       "FROM P_BOM_DETAIL D "
                       "JOIN B_ITEM I ON D.CHILD_ITEM_CD = I.ITEM_CD "
                       f"WHERE D.CHILD_ITEM_CD = '{str(self.code)}'AND D.PRNT_PLANT_CD = 'P1'")

        row = self.cursor.fetchall()
        finding_models = deque(row)
        models = []

        # finding_models = [prnt_item_cd, child_item_cd, item_nm, spec, acct, valid_from_dt]
        while finding_models:
            model = finding_models.popleft()
            prnt_item_cd = model[0]
            child_item_cd = model[1]
            item_nm = model[2]
            spec = model[3]
            valid_from_dt = model[4]
            acct = model[5]

            if acct == '10':
                models.append([prnt_item_cd,item_nm,valid_from_dt])
                continue

            self.cursor.execute("SELECT D.PRNT_ITEM_CD, D.CHILD_ITEM_CD, I.ITEM_NM, I.SPEC, I.VALID_FROM_DT, I.ITEM_ACCT "
                           "FROM P_BOM_DETAIL D "
                           "JOIN B_ITEM I ON D.PRNT_ITEM_CD = I.ITEM_CD "
                           f"WHERE '{prnt_item_cd}' = D.CHILD_ITEM_CD")

            row = self.cursor.fetchone()
            while row:
                finding_models.append(row)
                row = self.cursor.fetchone()

        model_dic = {}
        for model in models:
            prnt_item_cd = model[0]
            item_nm = model[1]
            valid_from_dt = model[2]
            rev = item_nm[:item_nm.rfind("-REV")]
            if "단종" in item_nm or "불가" in item_nm or "사용중지" in item_nm:
                continue
            if rev not in model_dic:
                model_dic[rev] = [item_nm,int(valid_from_dt.strftime("%Y%m%d"))]
            else:
                if model_dic[rev][1] <= int(valid_from_dt.strftime("%Y%m%d")):
                    model_dic[rev] = [item_nm,int(valid_from_dt.strftime("%Y%m%d"))]

        model_dic = sorted(model_dic.values())
        model_dic = [i[0] for i in model_dic]

        return model_dic # [[item_nm]]


    # 모델 별 모든 원자재 [[child_item_cd, item_nm, spec, valid_from_dt]]
    def find_codes_by_model(self, model_name):
        cnt = 0
        self.model_name = model_name

        self.cursor.execute("SELECT D.PRNT_ITEM_CD, D.CHILD_ITEM_CD, I.ITEM_NM, I.SPEC, I.VALID_FROM_DT, I.ITEM_ACCT "
                       "FROM P_BOM_DETAIL D "
                       "JOIN B_ITEM I ON D.CHILD_ITEM_CD = I.ITEM_CD "
                       f"WHERE D.PRNT_ITEM_CD = '{str(self.model_name)}'AND D.PRNT_PLANT_CD = 'P1' "
                       "ORDER BY D.CHILD_ITEM_SEQ ASC")


        row = self.cursor.fetchall()
        finding_codes = []
        codes = []
        for i in row:
            heapq.heappush(finding_codes, (cnt,i))
            cnt += 1

        # [prnt_item_cd, child_item_cd, item_nm, spec, valid_from_dt, acct]
        while finding_codes:
            code = heapq.heappop(finding_codes)
            code = code[1]
            prnt_item_cd = code[0]
            child_item_cd = code[1]
            item_nm = code[2]
            spec = code[3]
            valid_from_dt = code[4]
            acct = code[5]


            self.cursor.execute("SELECT D.PRNT_ITEM_CD, D.CHILD_ITEM_CD, I.ITEM_NM, I.SPEC, I.VALID_FROM_DT, I.ITEM_ACCT "
                           "FROM P_BOM_DETAIL D "
                           "JOIN B_ITEM I ON D.CHILD_ITEM_CD = I.ITEM_CD "
                           f"WHERE '{child_item_cd}' = D.PRNT_ITEM_CD "
                           "ORDER BY D.CHILD_ITEM_SEQ ASC")

            row = self.cursor.fetchone()
            while row:
                if row[5] != '10' and row[5] != '20': # row[5] == acct
                    codes.append([row[1],row[2],row[3],row[4]]) # child_item_cd, item_nm, spec, valid_from_dt
                    row = self.cursor.fetchone()
                    continue
                # finding_codes.appendleft(row)
                heapq.heappush(finding_codes,(cnt,row))
                cnt += 1
                row = self.cursor.fetchone()

        return codes # [[child_item_cd, item_nm, spec, valid_from_dt]]

    # 원자재 단가, 화폐 (본사db)
    def find_currency(self):

        self.cursor.execute("SELECT TOP 1 M.ITEM_CD, M.IV_PRC, H.IV_CUR FROM M_IV_DTL M "
                            "JOIN M_IV_HDR H ON M.IV_NO = H.IV_NO "
                            f"WHERE M.ITEM_CD LIKE '{str(self.code)}%' "
                            "ORDER BY M.INSRT_DT DESC")

        row = self.cursor.fetchone()
        # item_cd, 단가, 화폐

        return row

    # 원자재 단가, 화폐 (서룡db)
    def join_srerp_currency(self):
        conn = pymssql.connect(host=hq_host, user=id, password=pw, database='SRERP',
                               charset='utf8')
        cursor = conn.cursor()
        cursor.execute("SELECT TOP 1 M.ITEM_CD, M.IV_PRC, H.IV_CUR FROM M_IV_DTL M "
                       "JOIN M_IV_HDR H ON M.IV_NO = H.IV_NO "
                       f"WHERE M.ITEM_CD LIKE '{str(self.code)}%' "
                       "ORDER BY M.INSRT_DT DESC")
        row = cursor.fetchone()
        return row

    # 원자재 단가, 화폐 (위해db)
    def join_wherp_currecny(self):
        conn = pymssql.connect(host=wh_host, user=id, password=pw, database='WHERP',
                               charset='utf8')
        cursor = conn.cursor()
        cursor.execute("SELECT TOP 1 M.ITEM_CD, M.IV_PRC, H.IV_CUR FROM M_IV_DTL M "
                       "JOIN M_IV_HDR H ON M.IV_NO = H.IV_NO "
                       f"WHERE M.ITEM_CD LIKE '{str(self.code)}%' "
                       "ORDER BY M.INSRT_DT DESC")
        row = cursor.fetchone()
        return row

    @staticmethod
    def find_rate(nation):
        conn = pymssql.connect(host=hq_host, user=id, password=pw, database='HQERP',
                               charset='utf8')
        cursor = conn.cursor()
        cursor.execute("SELECT TOP 1 STD_RATE FROM b_daily_exchange_rate "
                       f"WHERE FROM_CURRENCY = '{nation}' ANd TO_CURRENCY = 'KRW' "
                       "ORDER BY APPRL_DT DESC")
        row = cursor.fetchone()
        return float(row[0])



class Client(Code):

    conn = pymssql.connect(host=hq_host, user=id, password=pw, database='HQERP', charset='utf8')
    cursor = conn.cursor()

    client_list = {}

    client_list["롯데기공"] = 110042
    client_list["위닉스"] = 104723
    client_list["경동나비엔"] = 110887
    client_list["NCM"] = 111034
    client_list["SK매직"] = 111566
    client_list["글로우원"] = 106480
    client_list["지티코"] = 108491
    client_list["두성하이텍"] = 110883
    client_list["승보일렉콤"] = 111871

    except_text_list = ["단종","불가"]
    model_limit_count = 1000

    lwb_client = load_workbook("P1_제품_품목정보조회_v2.2.xlsx")
    lws_client = lwb_client.active

    # 고객사 초기화
    def __init__(self, company):
        self.company = company

    # 고객사 모든 모델 # return 모델명
    def find_client_models(self):

        client_models = []

        # 고객사 있는 테이블에서 품목, 품목명, 규격
        for row in self.lws_client.iter_rows(min_row=5):
            if row[7].value == self.company:
                client_models.append(row[0].value)

        return client_models # return 모델명

    # 고객사 모델별 최신 리비전 모든 모델 # return 모델명
    def find_client_models_recently_revision(self):

        client_models = {}

        for row in self.lws_client.iter_rows(min_row=5):

            if row[7].value == self.company:
                if row[8].value != "단종" and row[8].value != "사용 불가":
                    rev = row[1].value[:row[1].value.rfind("-REV")]
                    date = int(row[30].value.strftime("%Y%m%d"))
                    if rev not in client_models:
                        client_models[rev] = [row[0].value, row[1].value, date]
                    else:
                        if client_models[rev][2] <= date:
                            client_models[rev] = [row[0].value, row[1].value, date]

        client_models = sorted(client_models.items())
        client_models = [i[1][0] for i in client_models]

        return client_models # return 모델명

    # 고객사 모델별 최신 리비전 모든 모델 nm # return nm
    def find_client_nm_by_model_recently_revision(self):

        client_models = {}

        for row in self.lws_client.iter_rows(min_row=5):

            if row[7].value == self.company:
                if row[8].value != "단종" and row[8].value != "사용 불가":
                    rev = row[1].value[:row[1].value.rfind("-REV")]
                    date = int(row[30].value.strftime("%Y%m%d"))
                    if rev not in client_models:
                        client_models[rev] = [row[0].value, row[1].value, date]
                    else:
                        if client_models[rev][2] <= date:
                            client_models[rev] = [row[0].value, row[1].value, date]

        client_models = sorted(client_models.items())
        client_models = [i[1][1] for i in client_models]

        return client_models # return nm


    # 고객사 매출
    def find_sales_avg_parameter(self, avg):
        self.avg = avg


        usd_rate = Code.find_rate("USD")

        today = datetime.datetime.now()
        today_y = today.year
        today_m = today.month

        betwenn_date = []
        for i in [self.avg, 0]:
            avg_y = i // 12
            avg_m = i % 12

            target_y = today_y
            if avg_y > 0:
                target_y = today_y - avg_y

            if today_m > avg_m:
                target_m = today_m - avg_m
            else:
                target_y = target_y - 1
                target_m = (12 + today_m) - avg_m

            betwenn_date.append(f"{target_y}-{str(target_m).zfill(2)}")
        start = betwenn_date[0]
        end = betwenn_date[1]

        self.cursor.execute("SELECT PAYER, BILL_AMT, CUR, BILL_DT FROM S_BILL_HDR "
                            f"WHERE PAYER='{self.client_list[f'{self.company}']}' AND BILL_DT > '{start}-01' AND BILL_DT < '{end}-01' "
                            "ORDER BY BILL_DT DESC")

        rows = self.cursor.fetchall()
        sales_total_avg = 0
        for row in rows:
            if row[2] == "KRW":
                sales_total_avg += row[1] / usd_rate
                continue
            sales_total_avg += row[1]

        sales_total_avg = int(sales_total_avg / self.avg)
        return sales_total_avg # 월 평균(달러)

    # 재고 정보 (위해 서버 연결 끊겨서 한번에 가져옴)
    @staticmethod
    def find_modify_storage():
        conn = pymssql.connect(host=wh_host, user=id, password=pw, database='WHERP',
                               charset='utf8')
        cursor = conn.cursor()
        cursor.execute("SELECT ITEM_CD, SL_CD, GOOD_ON_HAND_QTY "
                       "FROM I_ONHAND_STOCK WHERE SL_CD = 'S100' OR SL_CD = 'S200' OR SL_CD = 'S400'")

        storage_dic = {}
        rows = cursor.fetchall()
        for row in rows:
            storage_dic[row[0]] = [row[1],row[2]]

        return storage_dic
