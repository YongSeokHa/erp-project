import pymssql
from openpyxl import load_workbook, Workbook
import time
from erp_query import Client, Code
from erp_data_to_graph import create_graph
from address import *

# 재고 정보 (위해 서버 연결 끊겨서 한번에 가져옴)
def find_modify_storage():
    conn = pymssql.connect(host=wh_host, user=id, password=pw, database='WHERP',
                           charset='utf8')
    cursor = conn.cursor()
    cursor.execute("SELECT ITEM_CD, SL_CD, GOOD_ON_HAND_QTY "
                   "FROM I_ONHAND_STOCK WHERE SL_CD = 'S100' OR SL_CD = 'S200' OR SL_CD = 'S400'")

    storage_dic = {}
    rows = cursor.fetchall()
    for row in rows:
        storage_dic[row[0]] = [row[1],row[2]]

    return storage_dic


# 데이터, 엑셀로 옮기기 (단가, 화폐, 재고 총량, 재고 총액)
def total_currency_rate():
    price = currency_row[1]
    if price != None:
        price = float(price)
    else:
        price = 0
    currency = currency_row[2]
    ws["M" + str(row_cnt)].value = price
    ws["N" + str(row_cnt)].value = currency

    if currency == "USD":
        rate = 1
        ws["O" + str(row_cnt)].value = rate
        ws["P" + str(row_cnt)].value = str("$") + str(round(total * price, 2))
        total_price += total * price

    elif currency == "KRW":
        rate = 1 / usd_rate
        ws["O" + str(row_cnt)].value = rate
        ws["P" + str(row_cnt)].value = str("$") + str(round(total * price * rate, 2))
        total_price += total * price * rate

    elif currency == "CNY":
        rate = usd_rate / cny_rate
        ws["O" + str(row_cnt)].value = rate
        ws["P" + str(row_cnt)].value = str("$") + str(round(total * price * rate, 2))
        total_price += total * price * rate


def create_excel(client_name):
    global total_price, currency_row, wb, ws, row_cnt, total, usd_rate, cny_rate
    wb = Workbook()
    ws = wb.active
    ws.append(["품목","품목명","규격","해당 고객사 모델","그 외 적용 모델","진료 S100_W","진료 S100_J","내료 S200","내료 S200_R","내수 S400_Z","내수 S400_L","총 재고수량","단가","통화","환율","총 재고 금액"])

    save_name = "01"+client_name+".xlsx"

    # 모델별 모든 코드
    code_list = {} # {code : [nm,spec]}

    # 객체
    client = Client(client_name)

    # 고객사 최신 모델
    model_list = client.find_client_models_recently_revision()
    print(model_list)
    # 모델찾고 (제외빼고)
    for model in model_list:
        # 코드 찾고
        for code in client.find_codes_by_model(model):
            # 중복 없이 코드 담고
            if code[0] not in code_list:
                code_list[code[0]] = code[1:-1]
                print(code)


    # 코드 정렬
    code_list = sorted(code_list.items())

    # 환율정보
    usd_rate = Client.find_rate("USD")
    cny_rate = Client.find_rate("CNY")

    # 재고정보 ITEM_CD, SL_CD, GOOD_ON_HAND_QTY
    storage_dic = Client.find_modify_storage()

    # 고객사 최근 모델의 nm
    client_nm_list = client.find_client_nm_by_model_recently_revision()


    total_data = len(code_list)


    total_price = 0 # 총 재고
    code_cnt = 0 # 총 코드 갯수
    row_cnt = 1 # 코드중 재고 0 빼고 진행 사항
    for i in code_list:
        code_cnt += 1
        print(code_cnt)

        # 코드 재고 저장 dict
        s_storage = {
            "S100": [0, 0],
            "S200": [0, 0],
            "S400": [0, 0]
        }
        
        # storage_dic = ITEM_CD : [SL_CD, GOOD_ON_HAND_QTY]
        for cdx, code_alpha in enumerate(["W","J","","R","Z","L"]):
            cdx = cdx % 2
            if i[0]+code_alpha in storage_dic:
                if storage_dic[i[0]+code_alpha][0] == "S100":
                    if cdx == 0:
                        s_storage["S100"][cdx] = s_storage["S100"][cdx] + int(storage_dic[i[0] + code_alpha][1])
                    elif cdx == 1:
                        s_storage["S100"][cdx] = s_storage["S100"][cdx] + int(storage_dic[i[0] + code_alpha][1])

                elif storage_dic[i[0]+code_alpha][0] == "S200":
                    if cdx == 0:
                        s_storage["S200"][cdx] = s_storage["S200"][cdx] + int(storage_dic[i[0] + code_alpha][1])
                    elif cdx == 1:
                        s_storage["S200"][cdx] = s_storage["S200"][cdx] + int(storage_dic[i[0] + code_alpha][1])

                elif storage_dic[i[0]+code_alpha][0] == "S400":
                    if cdx == 0:
                        s_storage["S400"][cdx] = s_storage["S400"][cdx] + int(storage_dic[i[0] + code_alpha][1])
                    elif cdx == 1:
                        s_storage["S400"][cdx] = s_storage["S400"][cdx] + int(storage_dic[i[0] + code_alpha][1])

        # 재고 0 제외
        v_total = 0
        for v in s_storage.values():
            for v_ in v:
                v_total += v_

        if v_total == 0:
            continue


        row_cnt += 1

        # 재고 total
        total = s_storage["S200"][0] + s_storage["S200"][1]
        target = Code(i[0])

        # 고객사 모델
        client_nm_str = ""

        # 코드별 모델
        model_str = ""
        for model in target.find_models_recently_revison():

            if model in client_nm_list:
                client_nm_str += model+"\n"
                continue
            model_str += model+"\n"

        # 데이터 -> 엑셀
        ws["A"+str(row_cnt)].value = i[0] # 코드
        ws["B"+str(row_cnt)].value = i[1][0] # 품목명
        ws["C"+str(row_cnt)].value = i[1][1] # 규격

        # 해당 고객사 모델 리비전 최신
        ws["D" + str(row_cnt)].value = client_nm_str
        # 코드 역전개 모델 리비전 최신
        ws["E"+str(row_cnt)].value = model_str

        ws["F" + str(row_cnt)].value = s_storage["S100"][0]
        ws["G" + str(row_cnt)].value = s_storage["S100"][1]
        ws["H" + str(row_cnt)].value = s_storage["S200"][0]
        ws["I" + str(row_cnt)].value = s_storage["S200"][1]
        ws["J" + str(row_cnt)].value = s_storage["S400"][0]
        ws["K" + str(row_cnt)].value = s_storage["S400"][1]

        ws["L" + str(row_cnt)].value = total # 총 재고

        # 환율 정보

        # 본사 db
        currency_row = target.find_currency()
        if currency_row != None:
            price = currency_row[1]
            total_currency_rate()

        # 서룡 db
        elif target.join_srerp_currency() != None:
            currency_row = target.join_srerp_currency()
            total_currency_rate()

        # 위해 db
        elif target.join_wherp_currecny():
            currency_row = target.join_wherp_currecny()
            total_currency_rate()

        print(row_cnt,"/",total_data, client_name)

        # 50개씩 저장
        if row_cnt % 51 == 0:
            wb.save(save_name)
            time.sleep(10)
            wb = load_workbook(save_name)
            ws = wb.active

    row_cnt += 1
    ws["O" + str(row_cnt)].value = "총 합계"
    ws["P" + str(row_cnt)].value = int(total_price)
    row_cnt += 1
    ws["O" + str(row_cnt)].value = "월 평균"
    sales = client.find_sales_avg_parameter(12)
    ws["P" + str(row_cnt)].value = sales
    print(len(storage_dic))
    wb.save(save_name)
    end = time.time()
    print(end - start,"초")
    total_price = 0
    total = 0
    row_cnt = 1
    currency_row = None
    return total_price, sales

client_list = []

client_list.append("승보일렉콤")
client_list.append("지티코")
client_list.append("NCM")
client_list.append("롯데기공")
client_list.append("위닉스")
client_list.append("경동나비엔")
client_list.append("SK매직")
client_list.append("글로우원")
client_list.append("두성하이텍")

for cla in client_list:
    cla_total,cla_sales = create_excel(cla)
    create_graph(cla,cla_total,cla_sales)

